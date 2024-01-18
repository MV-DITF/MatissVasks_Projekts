from openpyxl import Workbook, load_workbook 
wb=load_workbook('Budzets_2023.xlsx')
ws=wb['11-Novembris']
max_row=ws.max_row
s=[]
j=[]
nsumma = 0
for row in range(7,max_row+1):
    nov=(ws['C' + str(row)].value)
    if nov:
        s.append(nov)
        nsumma = nsumma + nov

ws=wb['12-Decembris']
max_row=ws.max_row
s1=[]
j1=[]
dsumma = 0
for row in range(7,max_row+1):
    dec=(ws['C' + str(row)].value)
    if dec:
        s1.append(dec)
        dsumma = dsumma + dec

kop_summa = dsumma + nsumma
avg = kop_summa / 2
gada = avg * 12

print("Novembra kopējie patēriņi pārtikā", round(nsumma, 2))
print("Decembra kopējie patēriņi pārtikā", round(dsumma, 2))
print("Cik kopā pa abiem mēnešiem ir patērēta nauda pārtikā", kop_summa)
print("Vidējais patēriņš uz pārtiku mēnesī", round(avg, 2))
print("Aptuvenie patēriņi gadā uz pārtiku", round(gada, 2))