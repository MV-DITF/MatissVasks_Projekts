from openpyxl import Workbook, load_workbook 
wb=load_workbook('Budzets_2023.xlsx')

ws=wb['01-Janvaris']
max_row=ws.max_row
s1=[]
jsumma = 0
for row in range(7,max_row+1):
    jan=(ws['C' + str(row)].value)
    if jan:
        s1.append(jan)
        jsumma = jsumma + jan

ws=wb['02-Februaris']
max_row=ws.max_row
s2=[]
fsumma = 0
for row in range(7,max_row+1):
    feb=(ws['C' + str(row)].value)
    if feb:
        s2.append(feb)
        fsumma = fsumma + feb

ws=wb['03-Marts']
max_row=ws.max_row
s3=[]
msumma = 0
for row in range(7,max_row+1):
    mar=(ws['C' + str(row)].value)
    if mar:
        s3.append(mar)
        msumma = msumma + mar


ws=wb['04-Aprilis']
max_row=ws.max_row
s4=[]
asumma = 0
for row in range(7,max_row+1):
    apr=(ws['C' + str(row)].value)
    if apr:
        s4.append(apr)
        asumma = asumma + apr
        

ws=wb['05-Maijs']
max_row=ws.max_row
s5=[]
masumma = 0
for row in range(7,max_row+1):
    mai=(ws['C' + str(row)].value)
    if mai:
        s1.append(mai)
        masumma = masumma + mai

ws=wb['06-Junijs']
max_row=ws.max_row
s6=[]
junsumma = 0
for row in range(7,max_row+1):
    jun=(ws['C' + str(row)].value)
    if jun:
        s6.append(jun)
        junsumma = junsumma + jun

ws=wb['07-Julijs']
max_row=ws.max_row
s7=[]
julsumma = 0
for row in range(7,max_row+1):
    jul=(ws['C' + str(row)].value)
    if jul:
        s7.append(jul)
        julsumma = julsumma + jul

ws=wb['08-Augusts']
max_row=ws.max_row
s8=[]
augsumma = 0
for row in range(7,max_row+1):
    aug=(ws['C' + str(row)].value)
    if aug:
        s8.append(aug)
        augsumma = augsumma + aug

ws=wb['09-Septembris']
max_row=ws.max_row
s9=[]
ssumma = 0
for row in range(7,max_row+1):
    sep=(ws['C' + str(row)].value)
    if sep:
        s9.append(sep)
        ssumma = ssumma + sep

ws=wb['10-Oktobris']
max_row=ws.max_row
s10=[]
osumma = 0
for row in range(7,max_row+1):
    okt=(ws['C' + str(row)].value)
    if okt:
        s10.append(okt)
        osumma = osumma + okt

ws=wb['11-Novembris']
max_row=ws.max_row
s11=[]
nsumma = 0
for row in range(7,max_row+1):
    nov=(ws['C' + str(row)].value)
    if nov:
        s1.append(nov)
        nsumma = nsumma + nov

ws=wb['12-Decembris']
max_row=ws.max_row
s12=[]
dsumma = 0
for row in range(7,max_row+1):
    dec=(ws['C' + str(row)].value)
    if dec:
        s12.append(dec)
        dsumma = dsumma + dec

kop_summa = jsumma + fsumma + msumma + asumma + masumma + junsumma + julsumma + augsumma + ssumma + osumma + nsumma + dsumma
avg = kop_summa / 12
gada = avg * 12


print("Janvara kopējie patēriņi pārtikā", round(jsumma, 2))
print("Februara kopējie patēriņi pārtikā", round(fsumma, 2))
print("Marta kopējie patēriņi pārtikā", round(msumma, 2))
print("Aprīļa kopējie patēriņi pārtikā", round(asumma, 2))
print("Maija kopējie patēriņi pārtikā", round(masumma, 2))
print("Junija kopējie patēriņi pārtikā", round(junsumma, 2))
print("Julija kopējie patēriņi pārtikā", round(julsumma, 2))
print("Augusta kopējie patēriņi pārtikā", round(augsumma, 2))
print("Septembra kopējie patēriņi pārtikā", round(ssumma, 2))
print("Oktobra kopējie patēriņi pārtikā", round(osumma, 2))
print("Novembra kopējie patēriņi pārtikā", round(nsumma, 2))
print("Decembra kopējie patēriņi pārtikā", round(dsumma, 2))

print("Visa gada kopējie patēriņi pārtikā", round(kop_summa, 2))
print("Vidējais patēriņš uz pārtiku mēnesī", round(avg, 2))
print("Aptuvenie patēriņi gadā uz pārtiku", round(gada, 2))